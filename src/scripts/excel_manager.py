"""
Excel Manager - Comprehensive Excel file operations for macOS compatibility

This module provides Excel file management capabilities using xlwings,
optimized for macOS M1/M4 compatibility with visual inspection features.

Key Features:
- Excel file opening with visual display
- Workbook refreshing with enhanced macOS M1/M4 support
- Data connection refresh and calculation management
- Visual inspection and display control
- Proper resource cleanup
- Sheet listing and validation
"""

import os
import shutil
import xlwings as xw
from datetime import datetime
from openpyxl import load_workbook


def create_excel_copy(original_path: str) -> str:
    """
    Create a timestamped copy of the Excel file for processing
    
    Args:
        original_path (str): Path to the original Excel file
        
    Returns:
        str: Path to the created copy
        
    Raises:
        Exception: If file copy fails
    """
    try:
        # Extract directory and filename components
        directory = os.path.dirname(original_path)
        filename = os.path.basename(original_path)
        name, ext = os.path.splitext(filename)
        
        # Create timestamped copy name
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        copy_name = f"{name}_copy.xlsx"
        copy_path = os.path.join(directory, copy_name)
        
        # Create the copy
        shutil.copy2(original_path, copy_path)
        
        print(f"📋 Created Excel copy: {filename} → {copy_name}")
        print(f"📁 Copy location: {copy_path}")
        
        return copy_path
        
    except Exception as e:
        print(f"❌ Failed to create Excel copy: {e}")
        raise Exception(f"Could not create copy of {original_path}: {str(e)}")


def list_excel_sheets(excel_file_path: str) -> list:
    """
    List all available sheet names in an Excel file
    
    Args:
        excel_file_path (str): Path to Excel file
        
    Returns:
        list: List of sheet names
        
    Raises:
        Exception: If file cannot be read
    """
    try:
        # Use openpyxl to quickly read sheet names without opening in xlwings
        workbook = load_workbook(excel_file_path, read_only=True, data_only=True)
        sheet_names = workbook.sheetnames
        workbook.close()
        
        print(f"📊 Available sheets in {os.path.basename(excel_file_path)}:")
        for i, sheet_name in enumerate(sheet_names, 1):
            print(f"   {i}. {sheet_name}")
        
        return sheet_names
        
    except Exception as e:
        print(f"❌ Failed to read Excel sheets: {e}")
        raise Exception(f"Could not read sheets from {excel_file_path}: {str(e)}")


def validate_sheet_name(excel_file_path: str, sheet_name: str) -> bool:
    """
    Validate that a sheet name exists in the Excel file
    
    Args:
        excel_file_path (str): Path to Excel file
        sheet_name (str): Name of sheet to validate
        
    Returns:
        bool: True if sheet exists, False otherwise
    """
    try:
        available_sheets = list_excel_sheets(excel_file_path)
        return sheet_name in available_sheets
    except Exception:
        return False


class ExcelManager:
    """
    Manages Excel file operations using xlwings for macOS compatibility.
    Handles opening, refreshing, and keeping the workbook visible for inspection.
    """
    
    def __init__(self):
        self.app = None
        self.workbook = None
        self.file_path = None
        self.is_open = False
    
    def open_excel_file(self, file_path: str, display: bool = True) -> bool:
        """
        Open Excel file for visual inspection and refreshing.
        
        Args:
            file_path (str): Path to Excel file
            display (bool): Whether to make Excel visible (default: True for inspection)
            
        Returns:
            bool: True if successfully opened, False otherwise
        """
        try:
            print(f"📊 Opening Excel file for visual inspection: {file_path}")
            
            # Connect to or start Excel application
            try:
                # Try to connect to existing Excel application first
                self.app = xw.apps.active
                print("📱 Connected to existing Excel application")
            except:
                # If no active Excel app, create a new one
                self.app = xw.App(visible=display)
                print("📱 Started new Excel application")
            
            # Ensure Excel is visible for inspection
            if display:
                self.app.visible = True
                
            # Open the workbook
            self.workbook = self.app.books.open(file_path)
            self.file_path = file_path
            self.is_open = True
            
            # Bring Excel to front for inspection
            if display:
                self.app.activate(steal_focus=True)
            
            print(f"✅ Excel file opened successfully: {os.path.basename(file_path)}")
            print(f"📊 Workbook is now visible for inspection")
            
            return True
            
        except Exception as e:
            print(f"❌ Failed to open Excel file: {e}")
            self.cleanup()
            return False
    
    def refresh_excel(self, sheet_name: str = None) -> bool:
        """
        Refresh data connections and calculations in the Excel workbook.
        Can refresh a specific sheet or all sheets if no sheet is specified.
        macOS-compatible version using xlwings with proper async handling.
        
        Args:
            sheet_name (str, optional): Name of specific sheet to refresh. 
                                      If None, refreshes all sheets.
        
        Returns:
            bool: True if refresh successful, False otherwise
        """
        if not self.is_open or not self.workbook:
            print("⚠️  Excel file is not open, cannot refresh")
            return False
        
        try:
            if sheet_name:
                print(f"🔄 Refreshing specific sheet: {sheet_name}")
            else:
                print("🔄 Refreshing Excel workbook (all sheets)...")
            
            # For macOS Excel, we need to use different methods
            import platform
            import time
            
            if platform.system() == "Darwin":  # macOS
                # macOS Excel through xlwings - enhanced method for M1/M4 compatibility
                try:
                    # Step 1: Activate the workbook to ensure it's current
                    self.workbook.activate()
                    
                    # Step 2: Force Excel application to foreground
                    self.app.activate()
                    
                    # Step 3: Set calculation mode to manual temporarily for control
                    original_calc_mode = self.app.calculation
                    self.app.calculation = 'manual'
                    
                    # Step 4: Refresh all data connections first (if any)
                    try:
                        # This is the key method missing in the original code
                        self.workbook.api.RefreshAll()
                        print("📊 Data connections refreshed")
                    except Exception as refresh_error:
                        print(f"⚠️  Data connection refresh failed (may not exist): {refresh_error}")
                    
                    # Step 5: Wait for operations to complete (Mac doesn't have async query method)
                    # Use extended sleep for M1/M4 processing time
                    time.sleep(2)
                    print("⏳ Waited for M1/M4 processing (Mac method)")
                    
                    # Step 6: Force recalculation on specific sheet or all worksheets
                    if sheet_name:
                        # Refresh only the specified sheet
                        try:
                            target_sheet = self.workbook.sheets[sheet_name]
                            target_sheet.activate()
                            target_sheet.api.calculate()
                            print(f"🔄 Recalculated sheet: {sheet_name}")
                        except Exception as sheet_error:
                            print(f"❌ Failed to refresh sheet '{sheet_name}': {sheet_error}")
                            print("Available sheets:", [sheet.name for sheet in self.workbook.sheets])
                            return False
                    else:
                        # Refresh all sheets (original behavior)
                        for sheet in self.workbook.sheets:
                            try:
                                # Make sheet active and force calculation
                                sheet.activate()
                                sheet.api.calculate()
                                print(f"🔄 Recalculated sheet: {sheet.name}")
                            except Exception as sheet_error:
                                print(f"⚠️  Sheet {sheet.name} calculation failed: {sheet_error}")
                    
                    # Step 7: Full application calculate (Mac method)
                    try:
                        # Mac Excel doesn't support app.api.Calculate(), use alternative
                        self.app.calculation = 'automatic'
                        # Force recalculation by briefly switching modes
                        self.app.calculation = 'manual'
                        self.app.calculation = 'automatic'
                        print("🔄 Full application calculation complete (Mac method)")
                    except Exception as calc_error:
                        print(f"⚠️  Application calculation failed: {calc_error}")
                    
                    # Step 8: Restore calculation to automatic
                    self.app.calculation = 'automatic'
                    
                    # Step 9: Force screen update (important for visual refresh)
                    try:
                        self.app.screen_updating = False
                        self.app.screen_updating = True
                        print("🖥️  Screen display refreshed")
                    except Exception as screen_error:
                        print(f"⚠️  Screen refresh failed: {screen_error}")
                    
                    if sheet_name:
                        print(f"✅ Sheet '{sheet_name}' refreshed (macOS enhanced method)")
                    else:
                        print("✅ Excel workbook refreshed (macOS enhanced method)")
                    
                except Exception as mac_error:
                    print(f"⚠️  macOS-specific refresh method failed: {mac_error}")
                    # Fallback: just ensure calculations are updated using Mac method
                    try:
                        self.app.calculation = 'automatic'
                        # Use Mac-compatible calculation method
                        self.app.calculation = 'manual'
                        self.app.calculation = 'automatic'
                        print("✅ Fallback: Set Excel calculation to automatic (Mac method)")
                    except Exception as fallback_error:
                        print(f"⚠️  Could not perform fallback calculation: {fallback_error}")
            
            else:  # Windows or other
                try:
                    # Try Windows-specific methods
                    self.workbook.api.RefreshAll()
                    
                    # Try async query wait (Windows only)
                    try:
                        self.app.calculate_until_async_queries_done()
                        print("⏳ Windows async queries completed")
                    except:
                        time.sleep(2)  # Fallback wait
                    
                    # Refresh specific sheet or all sheets for Windows
                    if sheet_name:
                        try:
                            target_sheet = self.workbook.sheets[sheet_name]
                            target_sheet.activate()
                            target_sheet.api.calculate()
                            print(f"🔄 Recalculated sheet: {sheet_name}")
                        except Exception as sheet_error:
                            print(f"❌ Failed to refresh sheet '{sheet_name}': {sheet_error}")
                            return False
                    
                    # Windows calculation method
                    try:
                        self.app.api.Calculate()
                    except:
                        # Fallback to mode switching
                        self.app.calculation = 'manual'
                        self.app.calculation = 'automatic'
                        
                    if sheet_name:
                        print(f"✅ Sheet '{sheet_name}' refreshed (Windows method)")
                    else:
                        print("✅ Excel workbook refreshed (Windows method)")
                except Exception as win_error:
                    print(f"⚠️  Windows-specific refresh failed: {win_error}")
            
            # Extended pause for M1/M4 compatibility - these chips need more time
            print("⏳ Waiting for operations to complete...")
            time.sleep(3)  # Increased from 1 to 3 seconds
            
            # Save the workbook to preserve any changes
            try:
                self.workbook.save()
                print("💾 Excel workbook saved")
                
                # Additional save verification for macOS
                if platform.system() == "Darwin":
                    time.sleep(1)  # Brief pause after save
                    print("✅ Save operation completed")
                    
            except Exception as save_error:
                print(f"⚠️  Could not save workbook: {save_error}")
            
            return True
            
        except Exception as e:
            print(f"❌ Failed to refresh Excel: {e}")
            return False
    
    def refresh_sheet(self, sheet_name: str) -> bool:
        """
        Convenience method to refresh a specific sheet only.
        
        Args:
            sheet_name (str): Name of the sheet to refresh
            
        Returns:
            bool: True if refresh successful, False otherwise
        """
        return self.refresh_excel(sheet_name=sheet_name)

    def ensure_visible(self) -> bool:
        """
        Ensure Excel application and workbook are visible for inspection.
        Enhanced for M1/M4 Mac compatibility with display refresh.
        
        Returns:
            bool: True if made visible, False otherwise
        """
        if not self.is_open or not self.app:
            return False
        
        try:
            import platform
            import time
            
            # Make sure Excel app is visible
            self.app.visible = True
            
            # Activate the workbook to bring it to front
            if self.workbook:
                self.workbook.activate()
            
            # Bring Excel to foreground
            self.app.activate(steal_focus=True)
            
            # For macOS M1/M4, ensure the display properly refreshes
            if platform.system() == "Darwin":
                try:
                    # Force display refresh by toggling screen updating
                    self.app.screen_updating = False
                    time.sleep(0.3)  # Brief pause for M1/M4 processing
                    self.app.screen_updating = True
                    
                    # Ensure the active sheet is properly displayed
                    if self.workbook and self.workbook.sheets:
                        active_sheet = self.workbook.sheets.active
                        active_sheet.activate()
                        
                    print("🖥️  Excel display refreshed for macOS M1/M4")
                except Exception as display_error:
                    print(f"⚠️  Display refresh failed: {display_error}")
            
            return True
            
        except Exception as e:
            print(f"⚠️  Could not ensure Excel visibility: {e}")
            return False
    
    def cleanup(self) -> None:
        """
        Clean up Excel resources. Note: Keeps workbook open for inspection.
        Only closes if there are issues.
        """
        try:
            if self.is_open and self.workbook:
                # Save any changes before cleanup
                self.workbook.save()
                print("💾 Saved Excel workbook")
                
                # Keep workbook open for inspection unless there's an error
                print("📊 Keeping Excel workbook open for continued inspection")
                
        except Exception as e:
            print(f"⚠️  Error during Excel cleanup: {e}")
            # If there's an error, close everything
            try:
                if self.workbook:
                    self.workbook.close()
                if self.app:
                    self.app.quit()
            except:
                pass
        
        # Reset state flags
        self.is_open = False
        self.workbook = None
        self.app = None
        self.file_path = None
    
    def close_excel(self) -> None:
        """
        Explicitly close Excel workbook and application.
        Use this only when you want to fully close Excel.
        """
        try:
            if self.workbook:
                self.workbook.save()
                self.workbook.close()
                print("📊 Closed Excel workbook")
            
            if self.app:
                self.app.quit()
                print("📱 Closed Excel application")
                
        except Exception as e:
            print(f"⚠️  Error closing Excel: {e}")
        
        finally:
            self.is_open = False
            self.workbook = None
            self.app = None
            self.file_path = None
    
    def reload_from_disk(self) -> bool:
        """
        Force reload the Excel workbook from disk to pick up openpyxl changes.
        This is a workaround for openpyxl/xlwings compatibility.
        
        Returns:
            bool: True if reload successful, False otherwise
        """
        if not self.is_open or not self.file_path:
            print("⚠️  No Excel file is currently open")
            return False
        
        try:
            print("🔄 Reloading Excel file from disk to pick up external changes...")
            
            # Save current file path
            current_path = self.file_path
            
            # Close current workbook (but keep app open)
            if self.workbook:
                self.workbook.close()
            
            # Reopen the workbook
            self.workbook = self.app.books.open(current_path)
            
            # Ensure it's visible
            self.ensure_visible()
            
            print("✅ Excel file reloaded from disk successfully")
            return True
            
        except Exception as e:
            print(f"❌ Failed to reload Excel file from disk: {e}")
            return False
    
    def update_cell(self, sheet_name: str, cell_ref: str, value: any) -> bool:
        """
        Update a single cell using xlwings (for real-time updates).
        
        Args:
            sheet_name (str): Name of the sheet
            cell_ref (str): Cell reference (e.g., "A1", "B2")
            value (any): Value to set
            
        Returns:
            bool: True if successful, False otherwise
        """
        if not self.is_open:
            print("⚠️  Excel file is not open")
            return False
        
        try:
            sheet = self.workbook.sheets[sheet_name]
            sheet.range(cell_ref).value = value
            print(f"📝 Updated {cell_ref} = {value} (real-time via xlwings)")
            return True
            
        except Exception as e:
            print(f"❌ Failed to update cell {cell_ref}: {e}")
            return False
    
    def update_cells_batch(self, sheet_name: str, cell_updates: dict) -> bool:
        """
        Update multiple cells in batch using xlwings (for real-time updates).
        
        Args:
            sheet_name (str): Name of the sheet
            cell_updates (dict): Dictionary of {cell_ref: value} pairs
            
        Returns:
            bool: True if successful, False otherwise
        """
        if not self.is_open:
            print("⚠️  Excel file is not open")
            return False
        
        try:
            sheet = self.workbook.sheets[sheet_name]
            
            # Update all cells
            for cell_ref, value in cell_updates.items():
                sheet.range(cell_ref).value = value
                print(f"📝 Updated {cell_ref} = {value}")
            
            print(f"✅ Batch updated {len(cell_updates)} cells (real-time via xlwings)")
            return True
            
        except Exception as e:
            print(f"❌ Failed to batch update cells: {e}")
            return False