from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd

def parse_sheet_xlsx_with_mapping(path, sheet_name=None):
    """
    Simplified Excel parser that provides DataFrame markdown and basic metadata
    """
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    # 1) Basic dimensions
    max_row, max_col = ws.max_row, ws.max_column

    # 2) Create coordinate mapping
    coordinate_map = {}
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            col_letter = get_column_letter(c)
            excel_coord = f"{col_letter}{r}"
            coordinate_map[(r, c)] = excel_coord

    # 3) Handle merged cells -> expand to top-left value
    merged_ranges = list(ws.merged_cells.ranges)
    merged_map = {}
    for mr in merged_ranges:
        min_row, min_col, max_row_m, max_col_m = mr.min_row, mr.min_col, mr.max_row, mr.max_col
        top_left = ws.cell(min_row, min_col).value
        for r in range(min_row, max_row_m + 1):
            for c in range(min_col, max_col_m + 1):
                merged_map[(r, c)] = top_left

    # 4) Build enhanced grid with coordinate information
    grid = []
    grid_with_coords = []
    
    for r in range(1, max_row + 1):
        row_vals = []
        row_with_coords = []
        
        for c in range(1, max_col + 1):
            v = ws.cell(r, c).value
            if (r, c) in merged_map:
                v = merged_map[(r, c)]
            
            excel_coord = coordinate_map[(r, c)]
            row_vals.append(v)
            row_with_coords.append({
                'value': v,
                'coordinate': excel_coord,
                'row': r,
                'col': c,
                'col_letter': get_column_letter(c)
            })
        
        grid.append(row_vals)
        grid_with_coords.append(row_with_coords)

    # 5) Create DataFrame with Excel-style references
    df = pd.DataFrame(grid)
    excel_columns = [get_column_letter(c) for c in range(1, max_col + 1)]
    df.columns = excel_columns
    df.index = range(1, max_row + 1)

    # 6) Generate DataFrame markdown with Excel-style references
    markdown = df.to_markdown(index=True)

    # 7) Create metadata with coordinate mapping and grid information
    metadata = {
        'sheet_info': {
            'name': ws.title,
            'rows': max_row,
            'cols': max_col,
            'merged_ranges': [str(mr) for mr in merged_ranges]
        },
        'coordinate_map': coordinate_map,  # (row,col) -> "A1" mapping
        'grid_with_coords': grid_with_coords  # Enhanced grid with cell metadata
    }

    result = {
        "df": df,  # Excel-style DataFrame with A,B,C columns and 1,2,3 rows
        "markdown": markdown,
        "metadata": metadata
    }

    return result

if __name__ == "__main__":
    print("=== Simplified Excel Parser ===")

    docs_dir = "docs/"

    # Convert the excel file to markdown
    file_name = "itus-banking-sample"
    result = parse_sheet_xlsx_with_mapping(docs_dir + file_name + ".xlsx")

    output_file = docs_dir + file_name + ".md"
    with open(output_file, "w", encoding="utf-8") as f:
        f.write(result["markdown"])
    
    print(f"Excel-style DataFrame markdown saved to {output_file}")
    print("Metadata:", result["metadata"])