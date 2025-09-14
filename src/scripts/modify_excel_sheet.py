"""
Tool 2: Modify Excel Sheet

This tool physically modifies Excel files to accommodate new data by adding columns or rows.
It preserves formatting and updates the state with the new table ranges.

Key Features:
- Uses openpyxl for Excel file modification
- Preserves existing formatting and structure  
- Adds appropriate headers for new columns/rows
- Updates state with modified table ranges
- Raises exceptions on failure to halt entire process
"""

import sys
import os
import re
from typing import Dict, Any, Tuple
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string


def normalize_period_for_database(display_period):
    """
    Convert any period format to database-compatible format
    Handles all common Excel display formats
    """
    if not display_period:
        return display_period
    
    # Remove common separators and spaces for pattern matching
    cleaned = re.sub(r'[^\w]', '', str(display_period).upper())
    
    # Quarter patterns (highest priority)
    if re.match(r'Q\d+\d{2}', cleaned):
        # Q325 -> Q3 FY25, Q226 -> Q2 FY26
        quarter = cleaned[0:2]
        year = cleaned[2:]
        return f"{quarter} FY{year}"
    
    # Quarter with space patterns (Q2 26, Q4 25)
    elif re.match(r'Q\d+\s*\d{2}', str(display_period).upper()):
        parts = re.findall(r'Q(\d+)\s*(\d{2})', str(display_period).upper())
        if parts:
            quarter, year = parts[0]
            return f"Q{quarter} FY{year}"
    
    # Quarter with FY patterns (Q2FY26, Q1 FY25)
    elif re.match(r'Q\d+\s*FY\s*\d{2}', str(display_period).upper()):
        parts = re.findall(r'Q(\d+)\s*FY\s*(\d{2})', str(display_period).upper())
        if parts:
            quarter, year = parts[0]
            return f"Q{quarter} FY{year}"
    
    # Financial year patterns (FY25, FY 25)
    elif re.match(r'FY\s*\d{2}', str(display_period).upper()):
        year = re.findall(r'FY\s*(\d{2})', str(display_period).upper())[0]
        return f"FY{year}"
    
    # Calendar year patterns (2024, CY24, CY 24)
    elif re.match(r'(CY\s*)?\d{4}', str(display_period).upper()):
        year = re.findall(r'(\d{4})', str(display_period))[0]
        return f"CY{year}"
    
    # Default: return as-is 
    return str(display_period)


def parse_excel_range(range_str: str) -> Tuple[int, int, int, int]:
    """
    Parse Excel range string into start/end column and row numbers
    
    Args:
        range_str (str): Excel range like "A5:D15"
        
    Returns:
        Tuple[int, int, int, int]: (start_col, start_row, end_col, end_row)
        
    Example:
        "A5:D15" -> (1, 5, 4, 15)
    """
    pattern = r'([A-Z]+)(\d+):([A-Z]+)(\d+)'
    match = re.match(pattern, range_str)
    
    if not match:
        raise ValueError(f"Invalid Excel range format: {range_str}")
    
    start_col_letter, start_row_str, end_col_letter, end_row_str = match.groups()
    
    start_col = column_index_from_string(start_col_letter)
    start_row = int(start_row_str)
    end_col = column_index_from_string(end_col_letter)
    end_row = int(end_row_str)
    
    return start_col, start_row, end_col, end_row


def format_period_for_display(normalized_period: str, existing_format_pattern: str = "") -> str:
    """
    ALWAYS use normalized format for consistency
    
    Args:
        normalized_period (str): Database format like "Q2 FY26"
        existing_format_pattern (str): Pattern from existing headers (ignored for consistency)
        
    Returns:
        str: Always returns normalized format for database consistency
    """
    # ALWAYS return normalized format for consistency
    # This ensures Excel headers match what the API expects
    return normalized_period


def modify_excel_sheet(
    excel_file_path: str,
    table_range: str,
    modification_type: str,
    target_period: str,
    position: str,
    state: Dict[str, Any],
    target_cell: str = None
) -> None:
    """
    Modify Excel file to accommodate new data by adding columns or rows.
    
    This tool directly modifies:
    - The Excel file on disk
    - Updates table range in state["identified_tables"]
    - state["processing_status"] = "excel_modified"
    
    Args:
        excel_file_path (str): Path to the Excel file to modify
        table_range (str): Current table range to modify (e.g., "A5:D15")  
        modification_type (str): Type of modification ("add_column", "add_row", "insert_column")
        target_period (str): Period to add (e.g., "Q2 FY26")
        position (str): Where to add ("after_last", "before_first", "at_column_E")
        state (Dict[str, Any]): Agent state object to modify directly
        
    Raises:
        Exception: If Excel modification fails (halts entire process)
    """
    try:
        print(f"🔧 Tool 2: Modifying Excel sheet...")
        print(f"📁 File: {excel_file_path}")
        print(f"📊 Table range: {table_range}")
        print(f"🔨 Modification: {modification_type}")
        print(f"📅 Target period: {target_period}")
        print(f"📍 Position: {position}")
        
        # Normalize target period for consistency
        normalized_target_period = normalize_period_for_database(target_period)
        print(f"📅 Normalized target period: {normalized_target_period}")
        
        # Load the Excel workbook
        workbook = load_workbook(excel_file_path)
        worksheet = workbook.active
        
        # Parse the current table range
        start_col, start_row, end_col, end_row = parse_excel_range(table_range)
        
        changes_made = []
        new_range = table_range  # Default to original range
        
        if modification_type == "add_column":
            new_range, changes = _add_column_to_table(
                worksheet, start_col, start_row, end_col, end_row, 
                normalized_target_period, position, state, target_cell
            )
            changes_made.extend(changes)
            
        elif modification_type == "add_row":
            new_range, changes = _add_row_to_table(
                worksheet, start_col, start_row, end_col, end_row,
                normalized_target_period, position, state
            )
            changes_made.extend(changes)
            
        elif modification_type == "insert_column":
            new_range, changes = _insert_column_in_table(
                worksheet, start_col, start_row, end_col, end_row,
                normalized_target_period, position, state, target_cell
            )
            changes_made.extend(changes)
            
        else:
            raise ValueError(f"Unsupported modification type: {modification_type}")
        
        # Save the modified workbook
        workbook.save(excel_file_path)
        
        # Update the table range in state
        _update_table_range_in_state(state, table_range, new_range)
        
        # CRITICAL: Update period mapping after adding columns
        if modification_type == "add_column" and target_cell:
            _update_period_mapping_in_state(state, table_range, target_cell, normalized_target_period)
            _update_sheet_period_mapping(state, target_cell, normalized_target_period)
        
        # DIRECTLY MODIFY STATE OBJECT
        state["processing_status"] = "excel_modified"
        
        print(f"✅ Tool 2 Complete: Excel file modified successfully")
        print(f"📊 Updated range: {table_range} → {new_range}")
        print(f"🔨 Changes made: {changes_made}")
        print(f"📊 Status: {state['processing_status']}")
        
    except Exception as e:
        error_msg = f"Error in modify_excel_sheet: {str(e)}"
        state["errors"].append(error_msg)
        state["processing_status"] = "error"
        print(f"❌ Tool 2 Error: {error_msg}")
        # Re-raise to halt entire process as per requirements
        raise Exception(error_msg)


def _add_column_to_table(
    worksheet, start_col: int, start_row: int, end_col: int, end_row: int,
    target_period: str, position: str, state: Dict[str, Any], target_cell: str = None
) -> Tuple[str, list]:
    """Add a new column to the table range"""
    
    # Determine insertion position
    if position == "after_last":
        insert_col = end_col + 1
    elif position == "before_first":
        insert_col = start_col
    else:
        # Try to parse specific position like "at_column_E"
        if "column_" in position:
            col_letter = position.split("_")[-1]
            insert_col = column_index_from_string(col_letter)
        else:
            insert_col = end_col + 1  # Default to after last
    
    # Insert the new column in Excel
    worksheet.insert_cols(insert_col)
    
    changes_made = [f"Inserted column at position {get_column_letter(insert_col)}"]
    
    # Set the header in the specific cell provided by LLM
    if target_cell:
        # Parse the target cell (e.g., "E2" -> column E, row 2)
        from openpyxl.utils import coordinate_to_tuple
        target_row, target_col = coordinate_to_tuple(target_cell)
        
        # Format the period for display based on existing headers in the same row
        existing_headers = []
        for col in range(start_col, end_col + 1):
            cell_value = worksheet.cell(target_row, col).value
            if cell_value and any(period_indicator in str(cell_value).upper() 
                                for period_indicator in ['Q', 'FY', 'CY']):
                existing_headers.append(str(cell_value))
        
        display_period = target_period
        if existing_headers:
            # Try to match existing format pattern
            display_period = format_period_for_display(target_period, existing_headers[0])
        
        # Set the header in the exact cell specified by LLM
        header_cell = worksheet.cell(target_row, target_col)
        header_cell.value = display_period
        changes_made.append(f"Set {target_cell} = '{display_period}'")
        
        print(f"📋 Placed header '{display_period}' in cell {target_cell}")
    else:
        print("⚠️  No target_cell specified by LLM - skipping header placement")
    
    # Update the table range  
    new_end_col = max(end_col + 1, insert_col) if position != "before_first" else end_col + 1
    new_start_col = start_col if position != "before_first" else start_col
    
    new_range = f"{get_column_letter(new_start_col)}{start_row}:{get_column_letter(new_end_col)}{end_row}"
    
    return new_range, changes_made


def _add_row_to_table(
    worksheet, start_col: int, start_row: int, end_col: int, end_row: int,
    target_period: str, position: str, state: Dict[str, Any]
) -> Tuple[str, list]:
    """Add a new row to the table range"""
    
    # Determine insertion position  
    if position == "after_last":
        insert_row = end_row + 1
    elif position == "before_first":
        insert_row = start_row
    else:
        insert_row = end_row + 1  # Default to after last
    
    # Insert the new row in Excel
    worksheet.insert_rows(insert_row)
    
    changes_made = [f"Inserted row at position {insert_row}"]
    
    # Add metric label in the first column of the new row
    metric_col = start_col  # Typically column A
    new_cell = worksheet.cell(insert_row, metric_col)
    new_cell.value = f"New Metric ({target_period})"  # Placeholder - will be replaced during cell mapping
    changes_made.append(f"Set {get_column_letter(metric_col)}{insert_row} = 'New Metric ({target_period})'")
    
    # Update the table range
    new_end_row = max(end_row + 1, insert_row) if position != "before_first" else end_row + 1
    new_start_row = start_row if position != "before_first" else start_row
    
    new_range = f"{get_column_letter(start_col)}{new_start_row}:{get_column_letter(end_col)}{new_end_row}"
    
    return new_range, changes_made


def _insert_column_in_table(
    worksheet, start_col: int, start_row: int, end_col: int, end_row: int,
    target_period: str, position: str, state: Dict[str, Any], target_cell: str = None
) -> Tuple[str, list]:
    """Insert a column within the existing table range"""
    
    # For insert operations, we place the column in the middle or at a specific position
    if position == "middle":
        insert_col = start_col + (end_col - start_col) // 2
    else:
        insert_col = end_col  # Default to end of range
    
    # Use the same logic as add_column but with different positioning  
    return _add_column_to_table(worksheet, start_col, start_row, end_col, end_row,
                               target_period, f"at_column_{get_column_letter(insert_col)}", state, target_cell)


def _detect_header_rows(worksheet, start_row: int, end_row: int) -> list:
    """
    Detect which rows contain headers (quarters/periods) ABOVE the data table
    
    Args:
        start_row: First row of data table (e.g., 5 for A5:D15)
        end_row: Last row of data table
    
    Returns:
        list: Row numbers containing period headers
    """
    header_rows = []
    
    # Look for headers ABOVE the data table, not within it
    # For table A5:D15, check rows 1-4 for headers
    search_start = max(1, start_row - 4)  # Go back up to 4 rows
    search_end = start_row - 1  # Stop before data starts
    
    print(f"🔍 Searching for headers above data table in rows {search_start} to {search_end}")
    
    # Check rows above the data table for period indicators
    for row in range(search_start, search_end + 1):
        row_values = []
        for col in range(1, min(worksheet.max_column + 1, 15)):  # Check first 15 columns
            cell_value = worksheet.cell(row, col).value
            if cell_value:
                row_values.append(str(cell_value).strip())
        
        # Look for period indicators in the row
        row_text = ' '.join(row_values).upper()
        period_indicators = ['Q1', 'Q2', 'Q3', 'Q4', 'FY', 'CY', 'QUARTER', 'PERIOD']
        has_period = any(indicator in row_text for indicator in period_indicators)
        
        print(f"   Row {row}: {row_values[:5]}{'...' if len(row_values) > 5 else ''} -> Period indicators: {has_period}")
        
        if has_period:
            header_rows.append(row)
    
    # If no header rows found, use intelligent defaults based on table structure
    if not header_rows:
        print(f"⚠️  No period headers found above data table, using default header placement")
        # For a table starting at row 5, headers are typically in row 1
        if start_row >= 5:
            header_rows.append(1)  # Standard financial template: quarter headers in row 1
        elif start_row >= 2:
            header_rows.append(start_row - 1)  # Header in row above data
        else:
            header_rows.append(1)  # Fallback to row 1
    
    print(f"📋 Final header rows for new column: {header_rows}")
    return header_rows


def _update_table_range_in_state(state: Dict[str, Any], old_range: str, new_range: str) -> None:
    """
    Update the table range in state["identified_tables"] after modification
    """
    identified_tables = state.get("identified_tables", [])
    
    for table in identified_tables:
        if table.get("range") == old_range:
            table["range"] = new_range
            print(f"📊 Updated table range in state: {old_range} → {new_range}")
            break


def _update_period_mapping_in_state(state: Dict[str, Any], table_range: str, target_cell: str, period: str):
    """
    Update the period mapping in the current table's global_items after adding a column
    
    Args:
        state: Agent state object
        table_range: Current table range (may be old before update)
        target_cell: Cell where period header was placed (e.g., "E2")
        period: Normalized period value (e.g., "Q2 FY26")
    """
    try:
        # Extract column letter from target_cell (e.g., "E2" -> "E")
        import re
        col_match = re.match(r'([A-Z]+)', target_cell)
        if not col_match:
            print(f"⚠️  Could not extract column from target_cell: {target_cell}")
            return
        
        column_letter = col_match.group(1)
        
        # Find the current table in identified_tables
        identified_tables = state.get("identified_tables", [])
        current_table_index = state.get("current_table_index", 0)
        
        if 0 <= current_table_index < len(identified_tables):
            current_table = identified_tables[current_table_index]
            
            # Update period_mapping in global_items
            global_items = current_table.get("global_items", {})
            if "period_mapping" not in global_items:
                global_items["period_mapping"] = {}
            
            # Add the new column and period
            global_items["period_mapping"][column_letter] = period
            
            print(f"📅 Updated period mapping: Column {column_letter} = '{period}'")
            print(f"📅 Full period mapping: {global_items['period_mapping']}")
        else:
            print(f"⚠️  Could not find current table at index {current_table_index}")
            
    except Exception as e:
        print(f"⚠️  Error updating period mapping: {e}")
        # Don't raise - this is not critical enough to halt the process


def _update_sheet_period_mapping(state: Dict[str, Any], target_cell: str, period: str):
    """
    Update the sheet-global period mapping when a new column is added
    
    Args:
        state: Agent state object
        target_cell: Cell where period header was placed (e.g., "E2")
        period: Normalized period value (e.g., "Q1 FY25")
    """
    try:
        # Extract column letter from target_cell (e.g., "E2" -> "E")
        import re
        col_match = re.match(r'([A-Z]+)', target_cell)
        if not col_match:
            print(f"⚠️  Could not extract column from target_cell: {target_cell}")
            return
        
        column_letter = col_match.group(1)
        
        # Initialize sheet-global mappings if not exist
        if "sheet_period_mapping" not in state:
            state["sheet_period_mapping"] = {}
        if "sheet_columns_added" not in state:
            state["sheet_columns_added"] = []
        
        # Update sheet-global period mapping
        state["sheet_period_mapping"][column_letter] = period
        
        # Track that this period was added during this session
        if period not in state["sheet_columns_added"]:
            state["sheet_columns_added"].append(period)
        
        print(f"🌍 Updated SHEET-GLOBAL period mapping: Column {column_letter} = '{period}'")
        print(f"🌍 Columns added this session: {state['sheet_columns_added']}")
        
    except Exception as e:
        print(f"⚠️  Failed to update sheet period mapping: {e}")


def preserve_cell_formatting(source_cell, target_cell):
    """
    Copy formatting from source cell to target cell
    """
    if source_cell.font:
        target_cell.font = source_cell.font
    if source_cell.fill:
        target_cell.fill = source_cell.fill
    if source_cell.border:
        target_cell.border = source_cell.border
    if source_cell.alignment:
        target_cell.alignment = source_cell.alignment
    if source_cell.number_format:
        target_cell.number_format = source_cell.number_format


if __name__ == "__main__":
    """
    Test the Excel modification tool
    """
    print("=== Testing Excel Modification Tool ===")
    
    # This test requires an actual Excel file
    test_file = "test_modify.xlsx"
    
    # Mock state object
    test_state = {
        "identified_tables": [
            {
                "range": "A1:D5",
                "description": "Test Table"
            }
        ],
        "errors": [],
        "warnings": []
    }
    
    print("Note: This test requires an actual Excel file to modify.")
    print("Create a test Excel file and update the file path to run this test.")
    
    # Uncomment below to test with actual file:
    # try:
    #     modify_excel_sheet(
    #         excel_file_path=test_file,
    #         table_range="A1:D5", 
    #         modification_type="add_column",
    #         target_period="Q2 FY26",
    #         position="after_last",
    #         state=test_state
    #     )
    #     print("✅ Test completed successfully!")
    # except Exception as e:
    #     print(f"❌ Test failed: {e}")
