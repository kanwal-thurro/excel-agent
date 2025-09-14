"""
Tool 3: Cell Mapping and Fill Current Table

This tool applies preserved table-range global context and cell-specific extraction to map 
cells and fill values for the current table range only. It calls the xl_fill_plugin API
and updates the Excel file with results.

Key Features:
- Uses preserved global context from table identification (never re-evaluated)
- Extracts cell-specific context (metric from rows, quarter from columns)
- Makes parallel API calls to xl_fill_plugin /get-values endpoint
- Updates Excel file with retrieved values
- Directly modifies state with processing results
- LLM-based match selection with optional "no_match" capability

Environment Variables:
- GET_BEST_5: Enable top 5 match retrieval for LLM selection (default: false)
- ALLOW_NO_MATCH: Allow LLM to reject all matches if unsuitable (default: false)
- ENABLE_HUMAN_INTERVENTION: Enable human approval before tool execution (default: false)
  Example .env file:
  GET_BEST_5=true
  ALLOW_NO_MATCH=true
  ENABLE_HUMAN_INTERVENTION=false
  BACKEND_API_KEY=your_api_key_here
"""

import sys
import os
import asyncio
import aiohttp
import json
import re
from typing import Dict, Any, List, Tuple
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.comments import Comment
from openpyxl.styles import Font
import pandas as pd
from dotenv import load_dotenv

# Add parent directory to path to import existing modules
sys.path.append(os.path.join(os.path.dirname(__file__), '..'))

# Define normalize_period_for_database locally to avoid circular import
def normalize_period_for_database(display_period):
    """
    Convert any period format to database-compatible format
    Handles all common Excel display formats
    """
    if not display_period:
        return display_period
    
    import re
    
    # Remove common separators and spaces for pattern matching
    cleaned = re.sub(r'[^\w]', '', str(display_period).upper())
    
    # Quarter patterns (highest priority)
    if re.match(r'Q\d+\d{2}', cleaned):
        # Q325 -> Q3 FY25, Q226 -> Q2 FY26
        quarter = cleaned[0:2]
        year = cleaned[2:]
        return f"{quarter} FY{year}"
    
    # Quarter with space patterns (Q2 26, Q4 25)
    elif re.match(r'Q\d+\s*\d{2}', str(display_period).upper()):
        parts = re.findall(r'Q(\d+)\s*(\d{2})', str(display_period).upper())
        if parts:
            quarter, year = parts[0]
            return f"Q{quarter} FY{year}"
    
    # Quarter with FY patterns (Q2FY26, Q1 FY25)
    elif re.match(r'Q\d+\s*FY\s*\d{2}', str(display_period).upper()):
        parts = re.findall(r'Q(\d+)\s*FY\s*(\d{2})', str(display_period).upper())
        if parts:
            quarter, year = parts[0]
            return f"Q{quarter} FY{year}"
    
    # Financial year patterns (FY25, FY 25)
    elif re.match(r'FY\s*\d{2}', str(display_period).upper()):
        year = re.findall(r'FY\s*(\d{2})', str(display_period).upper())[0]
        return f"FY{year}"
    
    # Calendar year patterns (2024, CY24, CY 24)
    elif re.match(r'(CY\s*)?\d{4}', str(display_period).upper()):
        year = re.findall(r'(\d{4})', str(display_period))[0]
        return f"CY{year}"
    
    # Default: return as-is 
    return str(display_period)

# Import existing functionality
from scripts.excel_to_markdown import parse_sheet_xlsx_with_mapping

# Import centralized prompts
from scripts.prompts import create_match_selection_system_prompt, create_match_selection_user_prompt

# Load environment variables
load_dotenv()

# Configuration
XL_FILL_PLUGIN_BASE_URL = "https://localhost:8000"  # xl_fill_plugin Docker container
API_KEY = os.getenv('BACKEND_API_KEY', 'your_api_key_here')
MAX_CONCURRENT_REQUESTS = 5
GET_BEST_5 = os.getenv('GET_BEST_5', 'false').lower() == 'true'  # Toggle for getting top 5 results
ALLOW_NO_MATCH = os.getenv('ALLOW_NO_MATCH', 'false').lower() == 'true'  # Toggle for allowing LLM to select "no_match"
# When ALLOW_NO_MATCH=True: LLM can reject all top 5 matches if fundamentally incompatible
# When ALLOW_NO_MATCH=False: LLM must select best available match from top 5 (legacy behavior)


def extract_table_as_dataframe(table_range: str, excel_data: str) -> pd.DataFrame:
    """
    Extract a specific table range from Excel markdown data as a pandas DataFrame
    maintaining exact Excel coordinates
    
    Args:
        table_range (str): Excel range like "A5:D15"
        excel_data (str): Excel data in markdown format
        
    Returns:
        pd.DataFrame: DataFrame with Excel-style column names and row indices
    """
    # Parse the range
    pattern = r'([A-Z]+)(\d+):([A-Z]+)(\d+)'
    match = re.match(pattern, table_range)
    if not match:
        raise ValueError(f"Invalid Excel range format: {table_range}")
    
    start_col_letter, start_row_str, end_col_letter, end_row_str = match.groups()
    start_col = column_index_from_string(start_col_letter)
    start_row = int(start_row_str)
    end_col = column_index_from_string(end_col_letter)
    end_row = int(end_row_str)
    
    # Parse the markdown table - simplified extraction
    lines = excel_data.strip().split('\n')
    
    # Find data rows and extract values
    extracted_data = []
    for row_num in range(start_row, end_row + 1):
        row_data = []
        for col_num in range(start_col, end_col + 1):
            # This is a simplified extraction - in practice, you'd parse the markdown more carefully
            # For now, we'll create a placeholder DataFrame structure
            row_data.append("")  # Placeholder value
        extracted_data.append(row_data)
    
    # Create DataFrame with Excel-style column names and row indices
    column_names = [get_column_letter(col) for col in range(start_col, end_col + 1)]
    row_indices = list(range(start_row, end_row + 1))
    
    df = pd.DataFrame(extracted_data, columns=column_names, index=row_indices)
    return df


def extract_metric_from_row(excel_file_path: str, row: int) -> str:
    """
    Extract metric name from a specific row by looking at the leftmost columns
    
    Args:
        excel_file_path (str): Path to Excel file
        row (int): Row number to extract metric from
        
    Returns:
        str: Extracted metric name
    """
    try:
        # Load workbook and get the active sheet
        workbook = load_workbook(excel_file_path, data_only=True)
        worksheet = workbook.active
        
        # Search leftmost columns for metric name
        for col in range(1, 5):  # Check first 4 columns
            cell_value = worksheet.cell(row, col).value
            if cell_value and isinstance(cell_value, str):
                cell_value = str(cell_value).strip()
                if cell_value and not cell_value.isspace():
                    # Clean up metric name
                    return cell_value
        
        return f"Unknown Metric (Row {row})"
        
    except Exception as e:
        print(f"⚠️ Warning: Could not extract metric from row {row}: {e}")
        return f"Unknown Metric (Row {row})"


def extract_quarter_from_column(excel_file_path: str, col: str, table_range: str) -> str:
    """
    Extract quarter/period from a specific column by looking at header rows
    
    Args:
        excel_file_path (str): Path to Excel file
        col (str): Column letter (e.g., "E")
        table_range (str): Table range to determine header area
        
    Returns:
        str: Extracted quarter/period
    """
    try:
        # Load workbook and get the active sheet
        workbook = load_workbook(excel_file_path, data_only=True)
        worksheet = workbook.active
        
        # Convert column letter to number
        col_num = column_index_from_string(col)
        
        # Parse table range to get start row
        pattern = r'([A-Z]+)(\d+):([A-Z]+)(\d+)'
        match = re.match(pattern, table_range)
        if match:
            start_row = int(match.group(2))
            
            # Check header rows above the table start
            for check_row in range(max(1, start_row - 3), start_row):
                cell_value = worksheet.cell(check_row, col_num).value
                if cell_value and isinstance(cell_value, str):
                    cell_value = str(cell_value).strip()
                    # Look for quarter patterns
                    if any(pattern in cell_value.upper() for pattern in ['Q1', 'Q2', 'Q3', 'Q4', 'FY', 'CY']):
                        return cell_value
        
        return f"Unknown Period (Col {col})"
        
    except Exception as e:
        print(f"⚠️ Warning: Could not extract quarter from column {col}: {e}")
        return f"Unknown Period (Col {col})"


def normalize_period_for_api(period: str) -> str:
    """
    Normalize period format for API calls
    
    Args:
        period (str): Period in various formats
        
    Returns:
        str: Normalized period format for database
    """
    period = period.strip().upper()
    
    # Quarter patterns
    patterns = [
        (r'Q(\d)\s*(\d{2})', r'Q\1 FY\2'),          # "Q3 25" -> "Q3 FY25"
        (r'Q(\d)FY(\d{2})', r'Q\1 FY\2'),          # "Q3FY25" -> "Q3 FY25"
        (r'Q(\d)\s*FY\s*(\d{2})', r'Q\1 FY\2'),    # "Q3 FY 25" -> "Q3 FY25"
        (r'FY\s*(\d{2})', r'FY\1'),                 # "FY 25" -> "FY25"
        (r'(\d{4})', r'CY\1'),                      # "2024" -> "CY2024"
        (r'CY\s*(\d{2})', r'CY20\1')                # "CY24" -> "CY2024"
    ]
    
    for pattern, replacement in patterns:
        if re.match(pattern, period):
            return re.sub(pattern, replacement, period)
    
    return period  # Return as-is if no pattern matches


async def call_xl_fill_api(cell_mapping: Dict[str, Any]) -> Dict[str, Any]:
    """
    Call the xl_fill_plugin /get-values API with a single cell mapping
    
    Args:
        cell_mapping (Dict[str, Any]): Cell mapping with context
        
    Returns:
        Dict[str, Any]: API response data
    """
    url = f"{XL_FILL_PLUGIN_BASE_URL}/get-values"
    
    # Prepare request payload
    payload = {
        "company_name": cell_mapping.get("company_name", ""),
        "entity": cell_mapping.get("entity", ""),
        "metric": cell_mapping.get("metric", ""),
        "metric_type": cell_mapping.get("metric_type", ""),
        "quarter": cell_mapping.get("quarter", ""),
        "get_best_5": GET_BEST_5  # Include the toggle for top 5 results
    }
    
    headers = {
        "X-API-Key": API_KEY,
        "Content-Type": "application/json"
    }
    
    try:
        # Create SSL context that ignores certificate verification for localhost
        import ssl
        ssl_context = ssl.create_default_context()
        ssl_context.check_hostname = False
        ssl_context.verify_mode = ssl.CERT_NONE
        
        connector = aiohttp.TCPConnector(ssl=ssl_context)
        
        async with aiohttp.ClientSession(connector=connector) as session:
            async with session.post(url, json=payload, headers=headers) as response:
                if response.status == 200:
                    result = await response.json()
                    return {"status": "success", "data": result}
                else:
                    error_text = await response.text()
                    return {"status": "error", "error": f"HTTP {response.status}: {error_text}"}
                    
    except Exception as e:
        return {"status": "error", "error": str(e)}


def llm_select_best_match(top_5_matches: List[Dict[str, Any]], cell_mapping: Dict[str, Any]) -> Dict[str, Any]:
    """
    Use LLM to select the best match from top 5 results based on context compatibility
    
    Args:
        top_5_matches (List[Dict[str, Any]]): Top 5 matches from API
        cell_mapping (Dict[str, Any]): Original cell mapping context
        
    Returns:
        Dict[str, Any]: Selected best match from the top 5
    """
    try:
        print(f"🤖 LLM analyzing {len(top_5_matches)} top matches for best selection...")
        
        # Import get_llm_response from parent agent module
        sys.path.append(os.path.join(os.path.dirname(__file__), '..', 'agent'))
        from agent import get_llm_response
        
        # Prepare matches for LLM analysis
        matches_summary = []
        for i, match in enumerate(top_5_matches):
            match_info = {
                "rank": match.get("rank", i + 1),
                "value": match.get("value", ""),
                "company_name": match.get("company_name", ""),
                "entity": match.get("entity", ""),
                "metric_type": match.get("metric_type", ""),
                "metric": match.get("metric", ""),
                "time_period": match.get("time_period", ""),
                "document_year": match.get("document_year", ""),
                "hybrid_score": match.get("hybrid_score", 0),
                "source_url": match.get("source_url", "")
            }
            matches_summary.append(match_info)
        
        # Create LLM prompt using centralized prompts
        system_prompt = create_match_selection_system_prompt(allow_no_match=ALLOW_NO_MATCH)
        
        user_prompt = create_match_selection_user_prompt(cell_mapping, matches_summary)
        
        # Make LLM call
        messages = [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt}
        ]
        
        llm_response_text = get_llm_response(
            messages=messages,
            temperature=0,  # Low temperature for consistent selection
            max_tokens=1000,
            json_format=True
        )
        
        llm_result = json.loads(llm_response_text)
        selected_rank = llm_result.get("selected_rank", 1)
        reasoning = llm_result.get("reasoning", "No reasoning provided")
        confidence = llm_result.get("confidence", 0.5)
        
        print(f"🧠 LLM selected: {selected_rank} with confidence {confidence:.2f}")
        print(f"💭 Reasoning: {reasoning}")
        
        # Handle "no_match" case
        if selected_rank == "no_match":
            if ALLOW_NO_MATCH:
                print("🚫 LLM determined no suitable match from top 5 results")
                return {
                    "value": "",
                    "llm_selected": True,
                    "llm_reasoning": reasoning,
                    "llm_confidence": confidence,
                    "no_match_selected": True,
                    "rank": "no_match"
                }
            else:
                print("⚠️ LLM tried to select 'no_match' but ALLOW_NO_MATCH=False, falling back to rank 1")
                selected_rank = 1  # Force fallback to rank 1
        
        # Return the selected match for numeric ranks
        selected_match = None
        for match in top_5_matches:
            if match.get("rank", 0) == selected_rank:
                selected_match = match
                break
        
        if not selected_match:
            print(f"⚠️ Invalid rank {selected_rank} selected, falling back to rank 1")
            selected_match = top_5_matches[0]
        
        # Add LLM selection metadata
        selected_match["llm_selected"] = True
        selected_match["llm_reasoning"] = reasoning
        selected_match["llm_confidence"] = confidence
        selected_match["original_rank"] = selected_match.get("rank", 1)
        
        return selected_match
        
    except Exception as e:
        print(f"❌ LLM selection failed: {e}, falling back to rank 1")
        # Fallback to first match if LLM fails
        return top_5_matches[0] if top_5_matches else {}


async def process_cell_mappings_parallel(cell_mappings: Dict[str, Any]) -> Dict[str, Any]:
    """
    Process multiple cell mappings in parallel with controlled concurrency
    
    Args:
        cell_mappings (Dict[str, Any]): Dictionary of cell references to mappings
        
    Returns:
        Dict[str, Any]: API results for each cell
    """
    semaphore = asyncio.Semaphore(MAX_CONCURRENT_REQUESTS)
    
    async def process_single_cell(cell_ref: str, mapping: Dict[str, Any]) -> Tuple[str, Dict[str, Any]]:
        async with semaphore:
            result = await call_xl_fill_api(mapping)
            return cell_ref, result
    
    # Create tasks for all cells
    tasks = [
        process_single_cell(cell_ref, mapping)
        for cell_ref, mapping in cell_mappings.items()
    ]
    
    # Execute in parallel
    results = await asyncio.gather(*tasks)
    
    # Convert to dictionary
    api_results = {}
    for cell_ref, result in results:
        api_results[cell_ref] = result
    
    return api_results


def cell_mapping_and_fill_current_table(
    excel_data: str,
    table_range: str,
    global_items: Dict[str, Any],
    target_period: str,
    operation_type: str,
    state: Dict[str, Any]
) -> None:
    """
    Apply global context and cell-specific mapping for the current table range only.
    
    This tool directly modifies:
    - state["table_processing_results"][table_range] = Results
    - state["total_cells_filled"] += cells_filled  
    - state["processed_tables"].append(table_range)
    - state["current_table_index"] += 1
    - state["processing_status"] = "next_table" or "complete"
    - Excel file with API results
    
    Args:
        excel_data (str): Current Excel data in markdown format
        table_range (str): Table range to process (e.g., "A5:E15")
        global_items (Dict[str, Any]): Preserved global context for this table
        target_period (str): Target period (normalized format)
        operation_type (str): Operation type ("add_column", "update_existing", etc.)
        state (Dict[str, Any]): Agent state object to modify directly
        
    Raises:
        Exception: If processing fails (halts entire process)
    """
    try:
        print(f"🎯 Tool 3: Cell mapping and filling for table {table_range}")
        print(f"🌐 Global items: {global_items}")
        print(f"📅 Target period: {target_period}")
        print(f"🔨 Operation: {operation_type}")
        
        excel_file_path = state.get("excel_file_path", "")
        if not excel_file_path:
            raise Exception("Excel file path not found in state")
        
        # Generate cell mappings based on operation type
        cell_mappings = {}
        
        if operation_type == "add_column":
            cell_mappings = _generate_add_column_mappings(
                excel_file_path, table_range, global_items, target_period
            )
        elif operation_type == "update_existing":
            cell_mappings = _generate_update_existing_mappings(
                excel_file_path, table_range, global_items, target_period
            )
        elif operation_type == "add_metrics":
            cell_mappings = _generate_add_metrics_mappings(
                excel_file_path, table_range, global_items, target_period
            )
        else:
            raise Exception(f"Unsupported operation type: {operation_type}")
        
        print(f"📋 Generated {len(cell_mappings)} cell mappings")
        
        if not cell_mappings:
            print("⚠️ No cells to process for this table")
            # Still mark as processed
            state["processed_tables"].append(table_range)
            state["current_table_index"] += 1
            return
        
        # Make parallel API calls
        print(f"🌐 Making {len(cell_mappings)} API calls...")
        api_results = asyncio.run(process_cell_mappings_parallel(cell_mappings))
        
        # Process API results
        processed_results = {}
        cells_filled = 0
        cells_failed = 0
        
        for cell_ref, api_result in api_results.items():
            if api_result["status"] == "success":
                api_data = api_result["data"]
                
                # Extract value from API response
                matched_values = api_data.get("matched_values", {})
                top_5_matches = api_data.get("top_5_matches", [])
                
                if matched_values:
                    # Determine best match based on whether we have top 5 results
                    if GET_BEST_5 and top_5_matches and len(top_5_matches) > 1:
                        print(f"🔍 Cell {cell_ref}: Using LLM to select from {len(top_5_matches)} matches")
                        
                        # Get cell mapping context for LLM analysis
                        cell_mapping_context = cell_mappings.get(cell_ref, {})
                        
                        # Use LLM to select best match from top 5
                        best_match = llm_select_best_match(top_5_matches, cell_mapping_context)
                        
                        # Check if LLM selected "no_match" (only possible when ALLOW_NO_MATCH=True)
                        if best_match.get("no_match_selected", False):
                            print(f"🚫 Cell {cell_ref}: LLM determined no suitable match from top 5 results")
                            processed_results[cell_ref] = {
                                "value": "",
                                "status": "llm_no_match",
                                "reason": "LLM determined no suitable match from available options",
                                "llm_selected": True,
                                "llm_reasoning": best_match.get("llm_reasoning", ""),
                                "llm_confidence": best_match.get("llm_confidence", 0),
                                "total_alternatives": len(top_5_matches),
                                "api_response": api_data
                            }
                            cells_failed += 1
                            continue
                        
                        # Update value from LLM-selected match
                        value = best_match.get("value", "")
                        
                        print(f"🧠 Cell {cell_ref}: LLM selected rank {best_match.get('original_rank', 'N/A')} "
                              f"(confidence: {best_match.get('llm_confidence', 0):.2f})")
                        
                    else:
                        # Original logic: use the first (best) match from matched_values
                        best_match = list(matched_values.values())[0]
                        value = best_match.get("value", "")
                        
                        if GET_BEST_5:
                            print(f"📊 Cell {cell_ref}: Using default best match (no top 5 alternatives)")
                        else:
                            print(f"📊 Cell {cell_ref}: Using best match (GET_BEST_5 disabled)")
                    
                    processed_results[cell_ref] = {
                        "value": value,
                        "status": "filled",
                        "confidence_score": api_data.get("match_scores", {}).get(list(matched_values.keys())[0], [(0, 0.0)])[0][1] if api_data.get("match_scores") else 0.0,
                        "source_url": best_match.get("source_url", ""),
                        "value_in": best_match.get("value_in", ""),
                        "units": best_match.get("units", ""),
                        "document_year": best_match.get("document_year", ""),
                        "api_response": api_data,
                        "llm_selected": best_match.get("llm_selected", False),
                        "llm_reasoning": best_match.get("llm_reasoning", ""),
                        "llm_confidence": best_match.get("llm_confidence", 0),
                        "original_rank": best_match.get("original_rank", 1),
                        "total_alternatives": len(top_5_matches) if top_5_matches else 1
                    }
                    cells_filled += 1
                else:
                    processed_results[cell_ref] = {
                        "value": "",
                        "status": "no_data",
                        "reason": "No matching data found in database"
                    }
                    cells_failed += 1
            else:
                processed_results[cell_ref] = {
                    "value": "",
                    "status": "error",
                    "reason": api_result.get("error", "Unknown API error")
                }
                cells_failed += 1
        
        # Update Excel file with results
        _update_excel_with_results(excel_file_path, processed_results, cell_mappings)
        
        # DIRECTLY MODIFY STATE OBJECT
        if "table_processing_results" not in state:
            state["table_processing_results"] = {}
        
        state["table_processing_results"][table_range] = {
            "cell_mappings": cell_mappings,
            "api_results": processed_results,
            "cells_filled": cells_filled,
            "cells_failed": cells_failed,
            "table_processing_complete": True
        }
        
        state["total_cells_filled"] = state.get("total_cells_filled", 0) + cells_filled
        state["processed_tables"].append(table_range)
        state["current_table_index"] = state.get("current_table_index", 0) + 1
        
        # Check if all tables are processed
        identified_tables = state.get("identified_tables", [])
        if state["current_table_index"] >= len(identified_tables):
            state["processing_status"] = "complete"
        else:
            state["processing_status"] = "next_table"
        
        print(f"✅ Tool 3 Complete: Filled {cells_filled} cells, failed {cells_failed} cells")
        print(f"📊 Total cells filled: {state['total_cells_filled']}")
        print(f"📋 Processed tables: {len(state['processed_tables'])}")
        print(f"📊 Status: {state['processing_status']}")
        
    except Exception as e:
        error_msg = f"Error in cell_mapping_and_fill_current_table: {str(e)}"
        state["errors"].append(error_msg)
        state["processing_status"] = "error"
        print(f"❌ Tool 3 Error: {error_msg}")
        # Re-raise to halt entire process as per requirements
        raise Exception(error_msg)


def _generate_add_column_mappings(
    excel_file_path: str, table_range: str, global_items: Dict[str, Any], target_period: str
) -> Dict[str, Any]:
    """Generate cell mappings for add_column operation using period mapping"""
    
    cell_mappings = {}
    
    # Parse table range to find the data area
    pattern = r'([A-Z]+)(\d+):([A-Z]+)(\d+)'
    match = re.match(pattern, table_range)
    if not match:
        return cell_mappings
    
    start_row = int(match.group(2))
    end_col_letter = match.group(3)
    end_row = int(match.group(4))
    
    # Get period mapping from global items
    period_mapping = global_items.get("period_mapping", {})
    
    # CRITICAL FIX: Find the column that contains the target period
    target_column = None
    normalized_target_period = normalize_period_for_database(target_period)
    
    print(f"🔍 Looking for target period '{target_period}' (normalized: '{normalized_target_period}') in period mapping: {period_mapping}")
    
    # Search for the target period in the period mapping
    for col, period in period_mapping.items():
        normalized_period = normalize_period_for_database(period)
        print(f"🔍 Checking column {col}: '{period}' → '{normalized_period}' (match: {normalized_period == normalized_target_period})")
        if normalized_period == normalized_target_period:
            target_column = col
            print(f"✅ Found target period in column {col}")
            break
    
    if not target_column:
        print(f"⚠️  Target period '{target_period}' not found in period mapping. Falling back to rightmost column {end_col_letter}")
        target_column = end_col_letter
    
    # Use the found target column
    column_period = period_mapping.get(target_column, target_period)
    print(f"📅 Using column {target_column} for period: {column_period}")
    
    # Generate mappings for data rows in the target column (FIXED: use target_column, not end_col_letter)
    for row in range(start_row, end_row + 1):
        cell_ref = f"{target_column}{row}"  # CRITICAL FIX: Use target_column instead of end_col_letter
        
        # Extract metric from this row
        metric = extract_metric_from_row(excel_file_path, row)
        
        # Skip header rows (they typically don't have meaningful metrics)
        if any(header_word in metric.upper() for header_word in ['METRIC', 'FINANCIAL', 'KEY', 'RATIO']):
            continue
        
        # Create cell mapping using period from period_mapping
        cell_mappings[cell_ref] = {
            "company_name": global_items.get("company_name", ""),
            "entity": global_items.get("entity", ""),
            "metric_type": global_items.get("metric_type", ""),
            "metric": metric,
            "quarter": normalize_period_for_api(column_period),  # Use column-specific period
            "source_info": {
                "metric_source": f"Row {row}",
                "quarter_source": f"Period mapping column {target_column}",  # FIXED: Use target_column
                "global_source": "Preserved Global Context"
            }
        }
    
    return cell_mappings


def _generate_update_existing_mappings(
    excel_file_path: str, table_range: str, global_items: Dict[str, Any], target_period: str
) -> Dict[str, Any]:
    """Generate cell mappings for update_existing operation using period mapping"""
    
    cell_mappings = {}
    
    # Get period mapping from global items
    period_mapping = global_items.get("period_mapping", {})
    
    pattern = r'([A-Z]+)(\d+):([A-Z]+)(\d+)'
    match = re.match(pattern, table_range)
    if not match:
        return cell_mappings
    
    start_col_letter = match.group(1)
    start_row = int(match.group(2))
    end_col_letter = match.group(3)
    end_row = int(match.group(4))
    
    # Find the column that matches the target period using period mapping
    target_col = None
    for col, period in period_mapping.items():
        if normalize_period_for_api(period) == normalize_period_for_api(target_period):
            target_col = col
            break
    
    # If no matching column found, use end column as fallback
    if not target_col:
        target_col = end_col_letter
        print(f"⚠️  No matching column found for {target_period} in period_mapping, using {target_col}")
    else:
        print(f"📅 Found matching column {target_col} for period {target_period}")
    
    # Generate mappings for data rows in the target column
    for row in range(start_row, end_row + 1):
        cell_ref = f"{target_col}{row}"
        
        # Extract metric from this row
        metric = extract_metric_from_row(excel_file_path, row)
        
        # Skip header rows
        if any(header_word in metric.upper() for header_word in ['METRIC', 'FINANCIAL', 'KEY', 'RATIO']):
            continue
        
        # Get period for this column from period mapping
        column_period = period_mapping.get(target_col, target_period)
        
        # Create cell mapping
        cell_mappings[cell_ref] = {
            "company_name": global_items.get("company_name", ""),
            "entity": global_items.get("entity", ""),
            "metric_type": global_items.get("metric_type", ""),
            "metric": metric,
            "quarter": normalize_period_for_api(column_period),
            "source_info": {
                "metric_source": f"Row {row}",
                "quarter_source": f"Period mapping column {target_col}",
                "global_source": "Preserved Global Context"
            }
        }
    
    return cell_mappings


def _generate_add_metrics_mappings(
    excel_file_path: str, table_range: str, global_items: Dict[str, Any], target_period: str
) -> Dict[str, Any]:
    """Generate cell mappings for add_metrics operation"""
    
    # For add_metrics, we'd need to identify which new rows were added
    # This is a simplified implementation
    cell_mappings = {}
    
    print("ℹ️ add_metrics operation not fully implemented in this version")
    
    return cell_mappings


def _update_excel_with_results(excel_file_path: str, processed_results: Dict[str, Any], cell_mappings: Dict[str, Any] = None) -> None:
    """Update Excel file with API results, comments, and hyperlinks"""
    
    try:
        # Load workbook
        workbook = load_workbook(excel_file_path)
        worksheet = workbook.active
        
        # Update cells with values, comments, and hyperlinks
        for cell_ref, result in processed_results.items():
            cell = worksheet[cell_ref]
            
            if result["status"] == "filled" and result.get("value"):
                # Set the cell value
                value = result["value"]
                try:
                    if isinstance(value, str) and value.replace('.', '').replace('-', '').isdigit():
                        cell.value = float(value)
                    else:
                        cell.value = value
                except:
                    cell.value = value  # Keep as string if conversion fails
                
                # Add hyperlink to source if available
                source_url = result.get("source_url", "")
                if source_url and source_url.startswith("http"):
                    cell.hyperlink = source_url
                    # Style the hyperlink (blue and underlined)
                    cell.font = Font(color="0000FF", underline="single")
                
                # Add comment with matched parameters
                if cell_mappings and cell_ref in cell_mappings:
                    mapping = cell_mappings[cell_ref]
                    comment_text = _create_filled_cell_comment(mapping, result)
                    comment = Comment(comment_text, "Thurro Agent")
                    comment.width = 400   # Width in points (~5.5 Excel columns) 
                    comment.height = 120  # Height in points (~6 Excel rows)
                    cell.comment = comment
                    print(f"📝 Updated {cell_ref} = {value} (with comment & hyperlink)")
                else:
                    print(f"📝 Updated {cell_ref} = {value}")
            
            elif result["status"] == "no_data":
                cell.value = "N/A"
                
                # Add comment explaining no data found
                if cell_mappings and cell_ref in cell_mappings:
                    mapping = cell_mappings[cell_ref]
                    comment_text = _create_no_data_comment(mapping)
                    comment = Comment(comment_text, "Thurro Agent")
                    comment.width = 400   # Width in points (~5.5 Excel columns) 
                    comment.height = 120  # Height in points (~6 Excel rows)
                    cell.comment = comment
                    print(f"❌ No data for {cell_ref} (with comment)")
                else:
                    print(f"❌ No data for {cell_ref}")
            
            elif result["status"] == "llm_no_match":
                cell.value = "NO MATCH"
                
                # Add comment explaining LLM determined no suitable match
                if cell_mappings and cell_ref in cell_mappings:
                    mapping = cell_mappings[cell_ref]
                    comment_text = _create_llm_no_match_comment(mapping, result)
                    comment = Comment(comment_text, "Thurro Agent")
                    comment.width = 400   # Width in points (~5.5 Excel columns) 
                    comment.height = 120  # Height in points (~6 Excel rows)
                    cell.comment = comment
                    print(f"🚫 LLM no match for {cell_ref} (with comment)")
                else:
                    print(f"🚫 LLM no match for {cell_ref}")
            
            elif result["status"] == "error":
                cell.value = "ERROR"
                
                # Add comment explaining the error
                if cell_mappings and cell_ref in cell_mappings:
                    mapping = cell_mappings[cell_ref]
                    comment_text = _create_error_comment(mapping, result.get("reason", "Unknown error"))
                    comment = Comment(comment_text, "Thurro Agent")
                    comment.width = 400   # Width in points (~5.5 Excel columns) 
                    comment.height = 120  # Height in points (~6 Excel rows)
                    cell.comment = comment
                    print(f"❌ Error for {cell_ref} (with comment)")
                else:
                    print(f"❌ Error for {cell_ref}")
        
        # Save the updated workbook
        workbook.save(excel_file_path)
        print(f"💾 Saved updated Excel file")
        
    except Exception as e:
        print(f"⚠️ Warning: Could not update Excel file: {e}")


def _create_filled_cell_comment(mapping: Dict[str, Any], result: Dict[str, Any]) -> str:
    """Create pipe-separated comment with matched values from API response
    Format: company_name | entity | metric_type | metric | time_period | document_year
    """
    
    # Extract matched values from API response first, then fall back to mapping
    api_response = result.get("api_response", {})
    matched_values_dict = api_response.get("matched_values", {})
    
    # Get the first matched value entry (best match)
    matched_data = {}
    if matched_values_dict:
        first_key = list(matched_values_dict.keys())[0]
        matched_data = matched_values_dict[first_key]
    
    # Extract fields in the required format: company_name | entity | metric_type | metric | time_period | document_year
    company_name = matched_data.get("company_name", mapping.get("company_name", "")) or "N/A"
    entity = matched_data.get("entity", mapping.get("entity", "")) or "N/A"
    metric_type = matched_data.get("metric_type", mapping.get("metric_type", "")) or "N/A"
    metric = matched_data.get("metric", mapping.get("metric", "")) or "N/A"
    time_period = matched_data.get("time_period", mapping.get("quarter", "")) or "N/A"
    document_year = matched_data.get("document_year", result.get("document_year", "")) or "N/A"
    
    # Create pipe-separated comment in the required format
    comment_parts = [
        company_name,
        entity,
        metric_type,
        metric,
        time_period,
        document_year
    ]
    
    return " | ".join(comment_parts)


def _create_no_data_comment(mapping: Dict[str, Any]) -> str:
    """Create pipe-separated comment for no data cells
    Format: company_name | entity | metric_type | metric | time_period | document_year
    """
    
    # Create pipe-separated comment with search parameters (since no matched data)
    comment_parts = [
        mapping.get("company_name", "") or "N/A",
        mapping.get("entity", "") or "N/A", 
        mapping.get("metric_type", "") or "N/A",
        mapping.get("metric", "") or "N/A",
        mapping.get("quarter", "") or "N/A",
        "N/A"  # document_year not available for no data cases
    ]
    
    return "NO DATA: " + " | ".join(comment_parts)


def _create_error_comment(mapping: Dict[str, Any], error_reason: str) -> str:
    """Create pipe-separated comment for error cells
    Format: company_name | entity | metric_type | metric | time_period | document_year
    """
    
    # Create pipe-separated comment with search parameters (since error occurred)
    comment_parts = [
        mapping.get("company_name", "") or "N/A",
        mapping.get("entity", "") or "N/A", 
        mapping.get("metric_type", "") or "N/A",
        mapping.get("metric", "") or "N/A",
        mapping.get("quarter", "") or "N/A",
        "N/A"  # document_year not available for error cases
    ]
    
    return f"ERROR ({error_reason}): " + " | ".join(comment_parts)


def _create_llm_no_match_comment(mapping: Dict[str, Any], result: Dict[str, Any]) -> str:
    """Create pipe-separated comment for LLM no match cells
    Format: company_name | entity | metric_type | metric | time_period | document_year
    """
    
    # Create pipe-separated comment with search parameters and LLM reasoning
    comment_parts = [
        mapping.get("company_name", "") or "N/A",
        mapping.get("entity", "") or "N/A", 
        mapping.get("metric_type", "") or "N/A",
        mapping.get("metric", "") or "N/A",
        mapping.get("quarter", "") or "N/A",
        "N/A"  # document_year not available for no match cases
    ]
    
    llm_reasoning = result.get("llm_reasoning", "No reasoning provided")
    total_alternatives = result.get("total_alternatives", 0)
    
    return f"LLM NO MATCH ({total_alternatives} alternatives checked): {llm_reasoning} | " + " | ".join(comment_parts)


if __name__ == "__main__":
    """
    Test the cell mapping and filling tool
    """
    print("=== Testing Cell Mapping and Fill Tool ===")
    
    # Mock test data
    sample_excel_data = "Mock Excel data in markdown format"
    sample_table_range = "A5:E15"
    sample_global_items = {
        "company_name": "HDFC Bank",
        "entity": "HDFC Bank",
        "metric_type": ""
    }
    sample_target_period = "Q2 FY26"
    
    # Mock state object
    test_state = {
        "excel_file_path": "test_file.xlsx",
        "identified_tables": [{"range": sample_table_range}],
        "current_table_index": 0,
        "processed_tables": [],
        "errors": [],
        "warnings": []
    }
    
    print("Note: This test requires:")
    print("1. A running xl_fill_plugin server at http://localhost:8000")
    print("2. Valid API_KEY environment variable")
    print("3. An actual Excel file to modify")
    print("Uncomment the test code below when these prerequisites are met.")
    
    # Uncomment below to test with actual setup:
    # try:
    #     cell_mapping_and_fill_current_table(
    #         excel_data=sample_excel_data,
    #         table_range=sample_table_range,
    #         global_items=sample_global_items,
    #         target_period=sample_target_period,
    #         operation_type="add_column",
    #         state=test_state
    #     )
    #     print("✅ Test completed successfully!")
    # except Exception as e:
    #     print(f"❌ Test failed: {e}")
